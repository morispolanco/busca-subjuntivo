import streamlit as st
import spacy
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io

# Cargar el modelo de spaCy para español
try:
    nlp = spacy.load("es_core_news_sm")
except OSError:
    st.error("El modelo de spaCy para español no está instalado. Ejecuta: python -m spacy download es_core_news_sm")
    st.stop()

# Lista de terminaciones típicas del subjuntivo en español
subjuntivo_terminaciones = [
    'ara', 'aras', 'áramos', 'aran', 'are', 'ares', 'áremos', 'aren',  # -ar
    'iera', 'ieras', 'iéramos', 'ieran', 'iere', 'ieres', 'iéremos', 'ieren',  # -er/-ir irregular
    'era', 'eras', 'éramos', 'eran', 'ere', 'eres', 'éremos', 'eren',  # -er
    'iera', 'ieras', 'iéramos', 'ieran', 'iere', 'ieres', 'iéremos', 'ieren',  # -ir
    'a', 'as', 'amos', 'an', 'e', 'es', 'emos', 'en',  # presente
    'se', 'ses', 'semos', 'sen'  # imperfecto
]

# Verbos irregulares comunes en subjuntivo
verbos_irregulares_subjuntivo = [
    'sea', 'seas', 'seamos', 'sean',
    'vaya', 'vayas', 'vayamos', 'vayan',
    'haya', 'hayas', 'hayamos', 'hayan',
    'esté', 'estés', 'estemos', 'estén',
    'dé', 'des', 'demos', 'den',
    'sepa', 'sepas', 'sepamos', 'sepan',
    'quepa', 'quepas', 'quepamos', 'quepan'
]

def es_subjuntivo(token):
    """
    Determina si un token es un verbo en modo subjuntivo
    """
    # Verificar si es verbo
    if token.pos_ != 'VERB' and token.pos_ != 'AUX':
        return False
    
    # Verificar por terminaciones
    texto = token.text.lower()
    if any(texto.endswith(terminacion) for terminacion in subjuntivo_terminaciones):
        return True
    
    # Verificar verbos irregulares
    if texto in verbos_irregulares_subjuntivo:
        return True
    
    # Verificar análisis morfológico de spaCy
    if 'Mood=Sub' in token.morph:
        return True
    
    return False

def analizar_texto(texto):
    """
    Analiza el texto y encuentra todos los verbos en subjuntivo
    """
    doc = nlp(texto)
    resultados = []
    
    for sentencia in doc.sents:
        for token in sentencia:
            if es_subjuntivo(token):
                # Encontrar la cláusula completa
                clausula = obtener_clausula(token, sentencia)
                
                resultados.append({
                    'Verbo': token.text,
                    'Lemma': token.lemma_,
                    'Tiempo': obtener_tiempo_verbal(token),
                    'Persona': obtener_persona(token),
                    'Oración': sentencia.text.strip(),
                    'Cláusula': clausula,
                    'Posición': f"{token.idx}-{token.idx + len(token.text)}"
                })
    
    return resultados

def obtener_clausula(verbo, sentencia):
    """
    Extrae la cláusula subordinada que contiene el verbo en subjuntivo
    """
    clausula = []
    encontrado = False
    
    for token in sentencia:
        # Buscar conectores subordinantes típicos
        if token.text.lower() in ['que', 'cuando', 'si', 'aunque', 'para que', 'a fin de que']:
            encontrado = True
        
        if encontrado:
            clausula.append(token.text)
            
        if token == verbo:
            break
    
    return ' '.join(clausula) if clausula else sentencia.text

def obtener_tiempo_verbal(token):
    """
    Determina el tiempo verbal del verbo
    """
    morph = token.morph
    if 'Tense=Pres' in morph:
        return 'Presente'
    elif 'Tense=Past' in morph:
        return 'Pretérito'
    elif 'Tense=Imp' in morph:
        return 'Imperfecto'
    elif 'Tense=Fut' in morph:
        return 'Futuro'
    else:
        return 'Indeterminado'

def obtener_persona(token):
    """
    Determina la persona del verbo
    """
    morph = token.morph
    if 'Person=1' in morph and 'Number=Sing' in morph:
        return '1ra singular'
    elif 'Person=2' in morph and 'Number=Sing' in morph:
        return '2da singular'
    elif 'Person=3' in morph and 'Number=Sing' in morph:
        return '3ra singular'
    elif 'Person=1' in morph and 'Number=Plur' in morph:
        return '1ra plural'
    elif 'Person=2' in morph and 'Number=Plur' in morph:
        return '2da plural'
    elif 'Person=3' in morph and 'Number=Plur' in morph:
        return '3ra plural'
    else:
        return 'Indeterminada'

def crear_excel(resultados):
    """
    Crea un archivo Excel con los resultados
    """
    if not resultados:
        return None
    
    df = pd.DataFrame(resultados)
    
    # Crear el archivo Excel con formato
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Subjuntivos', index=False)
        
        # Obtener la hoja de trabajo para aplicar formato
        workbook = writer.book
        worksheet = writer.sheets['Subjuntivos']
        
        # Aplicar formato a los encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
        
        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def main():
    st.set_page_config(
        page_title="Analizador de Subjuntivo Español",
        page_icon="📝",
        layout="wide"
    )
    
    st.title("🔍 Analizador de Modo Subjuntivo en Español")
    st.markdown("""
    Esta aplicación identifica y analiza todas las formas verbales en modo subjuntivo 
    en textos en español y genera un informe detallado en Excel.
    """)
    
    # Sidebar con información
    with st.sidebar:
        st.header("ℹ️ Información")
        st.markdown("""
        **Características:**
        - Identifica verbos en subjuntivo
        - Analiza tiempo y persona verbal
        - Extrae cláusulas completas
        - Genera informe en Excel
        
        **Ejemplos de subjuntivo:**
        - Es importante que **estudies**
        - Ojalá **llueva** mañana
        - Quiero que **vengas** pronto
        """)
    
    # Área de texto para entrada
    col1, col2 = st.columns([2, 1])
    
    with col1:
        texto = st.text_area(
            "Introduce el texto a analizar:",
            height=300,
            placeholder="Ejemplo: Es necesario que estudies más para el examen. Ojalá que tengas suerte en tu viaje..."
        )
    
    with col2:
        st.markdown("### 📊 Estadísticas")
        if texto:
            doc = nlp(texto)
            total_palabras = len([token for token in doc if not token.is_punct])
            total_oraciones = len(list(doc.sents))
            st.metric("Palabras", total_palabras)
            st.metric("Oraciones", total_oraciones)
        else:
            st.info("Introduce texto para ver estadísticas")
    
    # Botón para analizar
    if st.button("🔍 Analizar Subjuntivo", type="primary"):
        if not texto.strip():
            st.warning("Por favor, introduce un texto para analizar.")
            return
        
        with st.spinner("Analizando texto..."):
            resultados = analizar_texto(texto)
        
        if resultados:
            st.success(f"✅ Se encontraron {len(resultados)} verbos en subjuntivo")
            
            # Mostrar resultados en tabla
            st.subheader("📋 Resultados del Análisis")
            df = pd.DataFrame(resultados)
            st.dataframe(df, use_container_width=True)
            
            # Generar y descargar Excel
            excel_file = crear_excel(resultados)
            
            st.download_button(
                label="📥 Descargar Informe Excel",
                data=excel_file,
                file_name="analisis_subjuntivo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Mostrar estadísticas
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total subjuntivos", len(resultados))
            with col2:
                tiempos = df['Tiempo'].value_counts()
                st.metric("Tiempo más común", tiempos.index[0] if len(tiempos) > 0 else "N/A")
            with col3:
                st.metric("Verbos únicos", df['Lemma'].nunique())
            
        else:
            st.info("ℹ️ No se encontraron verbos en modo subjuntivo en el texto.")
    
    # Ejemplos predefinidos
    st.subheader("💡 Ejemplos para probar")
    ejemplos = {
        "Ejemplo 1": "Es importante que estudies para el examen. Ojalá que tengas buena suerte.",
        "Ejemplo 2": "Quiero que vengas a la fiesta. Dudo que ella pueda asistir.",
        "Ejemplo 3": "Sería bueno que lloviera pronto. Temo que se sequen las plantas."
    }
    
    for nombre, ejemplo in ejemplos.items():
        if st.button(f"📌 {nombre}"):
            st.session_state.texto = ejemplo
            st.rerun()

if __name__ == "__main__":
    main()
